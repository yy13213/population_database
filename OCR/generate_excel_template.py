# -*-coding:utf-8-*-
"""
生成Excel模板文件
用于批量导入人口数据
"""
import pandas as pd
from datetime import datetime

def generate_template():
    """生成空白Excel模板"""
    
    # 定义列名和说明
    columns = [
        '身份证号码',
        '姓名',
        '曾用名',
        '性别',
        '出生年月日',
        '民族',
        '婚姻状况',
        '受教育程度',
        '户籍所在地-省',
        '户籍所在地-市',
        '户籍所在地-区',
        '住房情况',
        '现居住地-省',
        '现居住地-市',
        '现居住地-区',
        '户籍登记类型',
        '收入情况(元/月)',
        '数据来源'
    ]
    
    # 创建空DataFrame
    df = pd.DataFrame(columns=columns)
    
    # 添加说明行（作为第一行数据）
    explanation = [
        '18位身份证号',
        '必填',
        '可选',
        '男/女',
        '格式: 2000-01-01',
        '例: 汉族',
        '例: 未婚/已婚/离异/丧偶',
        '例: 本科',
        '例: 北京市',
        '例: 市辖区',
        '例: 东城区',
        '例: 自有住房',
        '例: 上海市',
        '例: 黄浦区',
        '例: 某某街道',
        '例: 家庭户/集体户',
        '例: 8000.50',
        '例: excel/manual'
    ]
    
    df.loc[0] = explanation
    
    # 保存为Excel
    filename = 'population_template.xlsx'
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='人口信息表', index=False)
        
        # 获取工作表
        workbook = writer.book
        worksheet = writer.sheets['人口信息表']
        
        # 设置列宽
        column_widths = {
            'A': 20,  # 身份证号码
            'B': 12,  # 姓名
            'C': 12,  # 曾用名
            'D': 8,   # 性别
            'E': 15,  # 出生年月日
            'F': 10,  # 民族
            'G': 12,  # 婚姻状况
            'H': 15,  # 受教育程度
            'I': 12,  # 户籍省
            'J': 12,  # 户籍市
            'K': 12,  # 户籍区
            'L': 15,  # 住房情况
            'M': 12,  # 现居住省
            'N': 12,  # 现居住市
            'O': 15,  # 现居住区
            'P': 15,  # 户籍登记类型
            'Q': 15,  # 收入情况
            'R': 15,  # 数据来源
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 设置表头样式
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置说明行样式
        explanation_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        explanation_font = Font(italic=True, size=9)
        
        for cell in worksheet[2]:
            cell.fill = explanation_fill
            cell.font = explanation_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    print(f"✅ Excel模板已生成: {filename}")
    return filename

def generate_sample_data():
    """生成测试样例数据"""
    
    # 定义列名
    columns = [
        '身份证号码',
        '姓名',
        '曾用名',
        '性别',
        '出生年月日',
        '民族',
        '婚姻状况',
        '受教育程度',
        '户籍所在地-省',
        '户籍所在地-市',
        '户籍所在地-区',
        '住房情况',
        '现居住地-省',
        '现居住地-市',
        '现居住地-区',
        '户籍登记类型',
        '收入情况(元/月)',
        '数据来源'
    ]
    
    # 创建测试数据
    test_data = [
        [
            '110101199001011234',
            '张三',
            '',
            '男',
            '1990-01-01',
            '汉族',
            '已婚',
            '本科',
            '北京市',
            '市辖区',
            '东城区',
            '自有住房',
            '北京市',
            '市辖区',
            '朝阳区',
            '家庭户',
            12000.00,
            'excel'
        ],
        [
            '310101198505152345',
            '李四',
            '李小四',
            '女',
            '1985-05-15',
            '汉族',
            '已婚',
            '硕士及以上',
            '上海市',
            '市辖区',
            '黄浦区',
            '租赁住房',
            '上海市',
            '市辖区',
            '浦东新区',
            '家庭户',
            18000.50,
            'excel'
        ],
        [
            '440106199207203456',
            '王五',
            '',
            '男',
            '1992-07-20',
            '汉族',
            '未婚',
            '大专',
            '广东省',
            '广州市',
            '天河区',
            '租赁住房',
            '广东省',
            '广州市',
            '天河区',
            '集体户',
            9500.00,
            'excel'
        ],
        [
            '330106198812254567',
            '赵六',
            '',
            '女',
            '1988-12-25',
            '回族',
            '离异',
            '高中',
            '浙江省',
            '杭州市',
            '西湖区',
            '自有住房',
            '浙江省',
            '杭州市',
            '西湖区',
            '家庭户',
            7800.00,
            'excel'
        ],
        [
            '510107199503105678',
            '孙七',
            '',
            '男',
            '1995-03-10',
            '汉族',
            '未婚',
            '本科',
            '四川省',
            '成都市',
            '武侯区',
            '租赁住房',
            '四川省',
            '成都市',
            '武侯区',
            '家庭户',
            10500.00,
            'excel'
        ]
    ]
    
    # 创建DataFrame
    df = pd.DataFrame(test_data, columns=columns)
    
    # 保存为Excel
    filename = 'population_sample.xlsx'
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='人口信息表', index=False)
        
        # 获取工作表
        workbook = writer.book
        worksheet = writer.sheets['人口信息表']
        
        # 设置列宽
        column_widths = {
            'A': 20, 'B': 12, 'C': 12, 'D': 8, 'E': 15, 'F': 10,
            'G': 12, 'H': 15, 'I': 12, 'J': 12, 'K': 12, 'L': 15,
            'M': 12, 'N': 12, 'O': 15, 'P': 15, 'Q': 15, 'R': 15
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 设置表头样式
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    print(f"✅ 测试样例已生成: {filename}")
    return filename

if __name__ == '__main__':
    print("=" * 70)
    print("Excel模板和样例生成工具")
    print("=" * 70)
    
    print("\n1. 生成空白模板...")
    template_file = generate_template()
    
    print("\n2. 生成测试样例...")
    sample_file = generate_sample_data()
    
    print("\n" + "=" * 70)
    print("生成完成！")
    print(f"📄 空白模板: {template_file}")
    print(f"📄 测试样例: {sample_file}")
    print("=" * 70)

